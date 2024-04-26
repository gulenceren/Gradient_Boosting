# -*- coding: utf-8 -*-

#Import the necessary data science libraries
from openpyxl import load_workbook
import numpy as np
import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
from sklearn.metrics import confusion_matrix
from sklearn.metrics import f1_score
import lightgbm as lgb
from xgboost import XGBClassifier
from sklearn.ensemble import GradientBoostingClassifier
from catboost import CatBoostClassifier
import time
from warnings import filterwarnings
filterwarnings('ignore')
from openpyxl import load_workbook

# Convert 'M' to 0, 'B' to 1 in a specific column
def edit(dataframe, column_index, workbook):
    sheet = workbook.active
    values = dataframe.iloc[:, column_index].values
    row_num = 2  # Assuming data starts from row 2
    for value in values:
        rating = 0 if value == 'M' else (1 if value == 'B' else '0')
        sheet.cell(row=row_num, column=column_index + 1, value=rating)
        row_num += 1
    workbook.save('./newdata.xlsx')

df = pd.read_excel('./data.xlsx')
workbook = load_workbook('./data.xlsx')
edit(df, 31, workbook)

#%% 
tumveriseti=pd.read_excel(open('./newdata.xlsx','rb'))

veriseti=np.array(tumveriseti)
np.random.shuffle(veriseti)
X=veriseti[:,0:30]
y=veriseti[:,31]

y = y.astype(np.float64)
X=np.array(X)
X_train,X_test,y_train,y_test=train_test_split(X,y,test_size=0.3, random_state=42)

print("======================GradientBoosting==================== ")
clf = GradientBoostingClassifier()
start = time.time()
clf.fit(X_train, y_train)
y_pred = clf.predict(X_test)
gbm_time = time.time() - start
print('The passing time:', gbm_time)
gbm_acc=accuracy_score(y_test, y_pred)
print("Acc: ",gbm_acc)
f1=f1_score(y_test, y_pred,average='weighted')
print("F1-Score: ",f1)
print("============Confusion Matrix=========== ")
cm = confusion_matrix(y_test, y_pred)
print("Confusion Matrix: ",cm)

print("======================Xgboost==================== ")
clf = XGBClassifier()
start = time.time()
clf.fit(X_train, y_train)
y_pred = clf.predict(X_test)
xgb_time = time.time() - start
print('The passing time:', xgb_time)
xgb_acc=accuracy_score(y_test, y_pred)
print("Acc: ",xgb_acc)
f1=f1_score(y_test, y_pred,average='weighted')
print("F1-Score: ",f1)
print("============Confusion Matrix=========== ")
cm = confusion_matrix(y_test, y_pred)
print("Confusion Matrix: ",cm)

print("======================LightGbm==================== ")
clf = lgb.LGBMClassifier()
start = time.time()
clf.fit(X_train, y_train)
y_pred = clf.predict(X_test)
lgb_time = time.time() - start
print('The passing time:', lgb_time)
lgb_acc=accuracy_score(y_test, y_pred)
print("Acc: ",lgb_acc)
f1=f1_score(y_test, y_pred,average='weighted')
print("F1-Score: ",f1)
print("============Confusion Matrix=========== ")
cm = confusion_matrix(y_test, y_pred)
print("Confusion Matrix: ",cm)

print("======================CatBoost==================== ")
clf = CatBoostClassifier()
start = time.time()
clf.fit(X_train, y_train)
y_pred = clf.predict(X_test)
cbs_time = time.time() - start
print('The passing time:', cbs_time)
cbs_acc=accuracy_score(y_test, y_pred)
print("Acc: ",cbs_acc)
f1=f1_score(y_test, y_pred,average='weighted')
print("F1-Score: ",f1)
print("============Confusion Matrix=========== ")
cm = confusion_matrix(y_test, y_pred)
print("Confusion Matrix: ",cm)